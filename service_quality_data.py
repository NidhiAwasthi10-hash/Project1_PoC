import random
import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook

import matplotlib.pyplot as plt
import numpy as np


#function to generate random time (24 hours time format)

import random
from datetime import datetime, timedelta

def generate_random_time(bias_start_time, bias_end_time):
    bias_start = datetime.strptime(bias_start_time, "%H:%M").time()
    bias_end = datetime.strptime(bias_end_time, "%H:%M").time()

    # Decide if the time should be biased or not
    if random.random() < 0.8:  # 80% chance to pick from the biased range
        # Generate a time within the biased range
        target_time = (
            datetime.combine(datetime.today(), bias_start) + 
            timedelta(hours=random.uniform(0, (bias_end.hour - bias_start.hour)),
                      minutes=random.uniform(0, 59))
        ).time()
    else:  # 20% chance to pick from the rest of the day
        # Generate a time outside the biased range
        if random.random() < 0.5:  # 50% chance to choose before bias
            target_time = (
                datetime.combine(datetime.today(), datetime.strptime("00:00", "%H:%M").time()) + 
                timedelta(hours=random.uniform(0, bias_start.hour),
                          minutes=random.uniform(0, 59))
            ).time()
        else:  # 50% chance to choose after bias
            target_time = (
                datetime.combine(datetime.today(), bias_end) + 
                timedelta(hours=random.uniform(0, (24 - bias_end.hour)), 
                          minutes=random.uniform(0, 59))
            ).time()

    return target_time.strftime("%H:%M")



def generate_random_time_morning_or_evening():
    
    is_morning = random.choice([True, False])

    if is_morning:
        start_hour = "6:30"
        end_hour = "10:00"
    else:
        start_hour = "17:00"
        end_hour = "20:00"
    return generate_random_time(start_hour, end_hour)


#function to generate data for a specific service type

def generate_data(service_type):
    if service_type.lower() == "uhd video streaming":
        latency = random.uniform( 4, 20)
        jitter = random.uniform( 4.6, 5.9)
        bitrate = random.uniform( 5.7, 11)
        packet_loss = random.uniform( 0, 1)
        peak_data_rate_dl = random.uniform( 13, 20)
        peak_data_rate_ul = random.uniform( 3, 10)
        mobility = random.randint( 0, 500)
        service_reliability = random.uniform( 95, 100)
        availability = random.uniform( 99, 99.999)
        servival_time = random.uniform( 8, 16)
        experienced_data_rate_dl = random.randint( 500, 1000)
        experienced_data_rate_ul = random.randint( 100, 500)
        intruption_time = random.randint( 1000, 3000)
        time_of_access = generate_random_time("18:00", "22:00")

    elif service_type.lower() == "immersive experience":
        latency = random.uniform(7, 15)
        jitter = random.uniform(8, 21)
        bitrate = random.uniform(28, 50)
        packet_loss = random.uniform( 0, 5)
        peak_data_rate_dl = random.uniform( 12, 20)
        peak_data_rate_ul = random.uniform( 2, 10)
        mobility = random.randint( 0, 30)
        service_reliability = random.uniform( 95, 100)
        availability = random.uniform( 99.9, 100)
        servival_time = random.uniform( 1, 10)
        experienced_data_rate_dl = random.randint( 650, 1000)
        experienced_data_rate_ul = random.randint( 20, 50)
        intruption_time = random.uniform(0, 1.5)
        time_of_access = generate_random_time("15:00", "20:00")

    elif service_type.lower() == "smart grid":
        latency = random.randint( 5, 50)
        jitter = random.uniform( 0, 1)
        bitrate = random.uniform( 0, 1)
        packet_loss = random.uniform( 0, 0.0001)
        peak_data_rate_dl = random.uniform( 10, 20)
        peak_data_rate_ul = random.uniform(1.5, 10)
        mobility = random.randint( 0, 0)
        service_reliability = random.uniform( 99.99, 100)
        availability = random.uniform( 99.999, 99.9999)
        servival_time = random.uniform( 10, 25)
        experienced_data_rate_dl = random.uniform( 5, 10)
        experienced_data_rate_ul = random.uniform(5, 10)
        intruption_time = random.uniform( 0, 1.5)
        time_of_access = generate_random_time("8:00", "18:00")

    elif service_type.lower() == "e-health":
        latency = random.randint( 1, 10)
        jitter = random.uniform( 2, 11)
        bitrate = random.uniform( 10, 15)
        packet_loss = random.uniform( 0, 0.00000001)
        peak_data_rate_dl = random.uniform( 0.2, 0.4)
        peak_data_rate_ul = random.uniform(0.2, 0.3)
        mobility = random.randint( 0, 120)
        service_reliability = random.uniform( 99.9999, 100)
        availability = random.uniform( 99, 99.99999)
        servival_time = random.uniform( 1, 50)
        experienced_data_rate_dl = random.uniform( 10, 100)
        experienced_data_rate_ul = random.uniform(10, 100)
        intruption_time = random.uniform( 0, 1.1)
        time_of_access = generate_random_time("9:00", "17:00")

    elif service_type.lower() == "its":
        latency = random.randint( 10, 100)
        jitter = random.uniform( 15, 20)
        bitrate = random.uniform( 0.2, 0.53)
        packet_loss = random.uniform( 0, 0.1)
        peak_data_rate_dl = random.uniform( 12, 20)
        peak_data_rate_ul = random.uniform(2, 10)
        mobility = random.randint( 50, 500)
        service_reliability = random.uniform( 99.999, 100)
        availability = random.uniform( 99, 99.9999)
        servival_time = random.uniform( 85, 110)
        experienced_data_rate_dl = random.uniform( 1, 10)
        experienced_data_rate_ul = random.uniform(1, 10)
        intruption_time = random.uniform( 1000, 10000)
        time_of_access = generate_random_time_morning_or_evening()

    elif service_type.lower() == "vonr":
        latency = random.randint(20, 150)
        jitter = random.uniform(1, 30)
        bitrate = random.uniform(5, 10)
        packet_loss = random.uniform( 0, 1)
        peak_data_rate_dl = random.uniform( 11, 20)
        peak_data_rate_ul = random.uniform( 2, 10)
        mobility = random.randint( 0, 500)
        service_reliability = random.uniform( 99.9, 100)
        availability = random.uniform( 95, 99)
        servival_time = random.uniform(98, 105)
        experienced_data_rate_dl = random.uniform(28, 50)
        experienced_data_rate_ul = random.randint(10, 25)
        intruption_time = random.uniform(0, 1.5)
        time_of_access = generate_random_time("0:00", "23:00")

    elif service_type.lower() == "connected vehicles":
        latency = random.uniform( 3, 100)
        jitter = random.uniform( 0.3, 0.5)
        bitrate = random.uniform( 5, 10)
        packet_loss = random.uniform( 0, 0.001)
        peak_data_rate_dl = random.uniform( 0.7, 1.2)
        peak_data_rate_ul = random.uniform(0.015, 0.025)
        mobility = random.randint( 50, 250)
        service_reliability = random.uniform( 99.999, 100)
        availability = random.uniform( 95, 99)
        servival_time = random.uniform( 1, 50)
        experienced_data_rate_dl = random.uniform( 40, 50)
        experienced_data_rate_ul = random.uniform(20, 25)
        intruption_time = random.uniform( 0, 0.8)
        time_of_access = generate_random_time_morning_or_evening()

    elif service_type.lower() == "industry automation":
        latency = random.randint( 1, 50)
        jitter = random.uniform( 0.05, 0.11)
        bitrate = random.uniform( 0.5, 1)
        packet_loss = random.uniform( 0, 0.0000001)
        peak_data_rate_dl = random.uniform( 10, 20)
        peak_data_rate_ul = random.uniform(1, 10)
        mobility = random.randint( 0, 30)
        service_reliability = random.uniform( 99.999, 100)
        availability = random.uniform( 99.99 , 99.99999)
        servival_time = random.uniform( 0, 100)
        experienced_data_rate_dl = random.uniform( 1, 10)
        experienced_data_rate_ul = random.uniform(1, 10)
        intruption_time = random.randint( 0, 100)
        time_of_access = generate_random_time("8:00", "16:00")

    elif service_type.lower() == "video surveillance":
        latency = random.randint( 10, 20)
        jitter = random.uniform( 0.2, 2.1)
        bitrate = random.uniform( 2, 8)
        packet_loss = random.uniform( 0, 0.01)
        peak_data_rate_dl = random.uniform( 0.1, 2)
        peak_data_rate_ul = random.uniform(0.05, 0.1)
        mobility = random.randint( 0, 10)
        service_reliability = random.uniform( 99.999, 100)
        availability = random.uniform( 99, 99.999)
        servival_time = random.uniform( 1, 100)
        experienced_data_rate_dl = random.uniform( 50, 200)
        experienced_data_rate_ul = random.uniform(20, 100)
        intruption_time = random.uniform( 0, 1.5)
        time_of_access = generate_random_time("12:00", "6:00")




        #prepare data to export
    data = {
        "Service Type": [service_type],
        "Latency (ms)": [latency],
        "Jitter (ms)": [jitter],
        "Bitrate (Mbps)": [bitrate],
        "Packet Loss": [packet_loss],
        "Peak Data Rate DL (Gbps)": [peak_data_rate_dl],
        "Peak Data Rate UL (Gbps)": [peak_data_rate_ul],
        "Mobility (km/h)": [mobility],
        "Service reliability (%)": [service_reliability],
        "Availability (%)": [availability],
        "Servival Time (s)": [servival_time],
        "Experienced Data Rate DL (Mbps)": [experienced_data_rate_dl],
        "Experienced Data Rate UL (Mbps)": [experienced_data_rate_ul],
        "Intrruption Time (ms)": [intruption_time],
        "Time of Access": [time_of_access],
    }

    df = pd.DataFrame(data)
    return df

excel_filename = "service_quality_report.xlsx"
service_types = ["UHD video streaming", "Immersive Experience", "Smart Grid", "e-health", "ITS", "VoNR", "Connected Vehicles", "Industry Automation", "Video Surveillance"]


all_data = []

num  = int(input("Enter number of entries to be created for each service type" ))

for _ in range(num):
    for service_type in service_types:
        data = generate_data(service_type)
        if data is not None:
            all_data.append(data)
    
combined_df = pd.concat(all_data, ignore_index=True)


if os.path.exists(excel_filename):

    book = load_workbook(excel_filename)
    sheet = book.active


    next_row = sheet.max_row + 1

    for i, column in enumerate(combined_df.columns):
        for j, value in enumerate (combined_df[column]):
            sheet.cell(row=next_row + j, column=i + 1, value=value)

    book.save(excel_filename)

else:
    combined_df.to_excel(excel_filename, index=False)


print("Data has been successfully appended to the excel file.")
excel_filename = "service_quality_report.xlsx"

df = pd.read_excel(excel_filename)
print(df)

def time_to_minutes(time_str):
    time_obj = pd.to_datetime(time_str, format="%H:%M")
    return time_obj.hour * 60 + time_obj.minute

def show_line_graph():
    df['Time in Minutes'] = df ['Time of Access'].apply(time_to_minutes)
    
    df['Time Slot'] = (df['Time in Minutes'] // 60) * 60 # taking data for 2 hours time slot
    
    df['Midpoint'] = df['Time Slot'] + 30 # adding 1 hour to start of the time window
    
    service_types = df['Service Type'].unique()
    
    plt.figure(figsize=(8,6))
    
    for i, service_type in enumerate(service_types):
        service_data = df[df['Service Type'] == service_type]
    
        service_counts = service_data.groupby('Midpoint').size().astype(int)
    
        plt.plot(service_counts.index, service_counts.values,
                 label=service_type,
                 marker='o',
                 linestyle='-',
                 linewidth=2, alpha=0.7)
    
    plt.xlabel('Time Slot (in hours from Midnight)')
    plt.ylabel('Number of UEs')
    
    plt.title('Number of UEs in 1 Hr time interval')
    
    plt.legend(title="Service Type")
    
    #time_labels = ["{:01}:{:01}".format(i // 60, i% 60) for i in range (60, 1440, 60)]
    time_labels = ["{:02}:{:02}".format(i // 60, i% 60) for i in range (0, 1440, 60)]
    
    plt.xticks(ticks=np.arange(0, 1440, 60), labels=time_labels)
    
    plt.grid(True)
    
    plt.show()


import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import tkinter as tk
from tkinter import ttk

# Load the data from the Excel file
excel_filename = "service_quality_report.xlsx"
df = pd.read_excel(excel_filename)

# Function to show the pie chart
def show_pie_chart():
    service_counts = df['Service Type'].value_counts()

    plt.figure(figsize=(8, 6))
    plt.pie(service_counts, labels=service_counts.index, autopct=lambda p: f'{int(p * sum(service_counts) / 100)}', startangle=90)
    plt.axis('equal')
    plt.title('Number of UEs Accessing Each Service')
    plt.show()


# Create the main window
root = tk.Tk()
root.title("Service Quality Visualization")

# Create a dropdown menu to select the type of graph
graph_selection = ttk.Combobox(root, values=["Pie Chart", "Line Chart"])
graph_selection.set("Select a graph type")
graph_selection.pack(pady=10)

# Function to handle the selection and button click
def on_button_click():
    selected_graph = graph_selection.get()
    if selected_graph == "Pie Chart":
        show_pie_chart()
    elif selected_graph == "Line Chart":
        show_line_graph()  # Placeholder for line chart functionality
    else:
        print("Please select a valid graph.")

# Create a button to submit the selection
button = tk.Button(root, text="Show Graph", command=on_button_click)
button.pack(pady=10)

# Run the main loop
root.mainloop()
